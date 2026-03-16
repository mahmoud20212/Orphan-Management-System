import os
from datetime import date, datetime
from pathlib import Path
from jinja2 import Environment, FileSystemLoader
from decimal import Decimal
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from utils import calculate_age
from services.db_services import DBService
from database.models import (
    Currency,
    Deceased,
    DeceasedBalance,
    DeceasedTransaction,
    GuardianTransaction,
    OrphanGuardian,
    Transaction,
    TransactionTypeEnum,
)

# المكتبات المطلوبة لدعم العربية والـ PDF
try:
    from weasyprint import HTML
    RENDERER = "weasy"
except Exception:
    import pdfkit
    RENDERER = "pdfkit"

# مسار جذر المشروع
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(PROJECT_ROOT, "templates")
ASSETS_IMAGES = Path(os.path.join(PROJECT_ROOT, "assets", "images")).resolve()

env = Environment(loader=FileSystemLoader(TEMPLATE_PATH))

logo_path = (ASSETS_IMAGES / "logo.png").as_uri() if (ASSETS_IMAGES / "logo.png").exists() else None
ps_logo_path = (ASSETS_IMAGES / "ps_logo.png").as_uri() if (ASSETS_IMAGES / "ps_logo.png").exists() else None


def _format_date(d):
    if not d: return "---"
    if isinstance(d, (date, datetime)):
        return d.strftime("%Y/%m/%d")
    return str(d)


def _get_attr(obj, name, default=None):
    if obj is None: return default
    return getattr(obj, name, default)


def _build_full_balances(balances, all_currencies):
    """Return list of balances for all currencies (include 0.0 for missing)."""
    bal_map = {}
    for b in balances:
        cur = getattr(b, 'currency', None)
        cid = getattr(cur, 'id', None) if cur is not None else getattr(b, 'currency_id', None)
        try:
            bal_map[int(cid)] = float(b.balance)
        except Exception:
            # ignore entries without usable currency id or balance
            continue

    result = []
    for curr in all_currencies:
        amt = float(bal_map.get(curr.id, 0.0))
        name = getattr(curr, 'name', None) or getattr(curr, 'code', str(getattr(curr, 'id', '')))
        result.append({"currency": name, "amount": amt})
    return result


def fetch_entity_data(entity_type: str, entity_id: int, db_service: DBService, user):
    common = {
        "generated_at": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "organization": "مؤسسة إدارة وتنمية أموال اليتامى",
        "exported_by": user.name,
        "logo_path": logo_path,
        "ps_logo_path": ps_logo_path,
    }

    if entity_type == "orphan":
        orphan = db_service.get_orphan_details(entity_id)
        if not orphan: return None

        gender_display = "غير محدد"
        if orphan.gender:
            if orphan.gender.name == "male": gender_display = "ذكر"
            elif orphan.gender.name == "female": gender_display = "أنثى"

        deceased_obj = orphan.deceased
        deceased_info = {
            "name": _get_attr(deceased_obj, 'name', '---'),
            "national_id": _get_attr(deceased_obj, 'national_id', '---'),
            "death_date": _format_date(_get_attr(deceased_obj, 'date_of_death')),
            "account_number": _get_attr(orphan, 'account_number', '---'),
            "archives_number": _get_attr(orphan, 'archives_number', '---'),
        }

        all_currencies = db_service.get_currencies()
        balances = _build_full_balances(orphan.balances, all_currencies)
        guardian_history = []
        primary_guardian = None
        sorted_links = sorted(orphan.guardian_links, key=lambda x: x.start_date or date.min, reverse=True)

        for link in sorted_links:
            g_data = {
                "name": link.guardian.name,
                "national_id": link.guardian.national_id or "---",
                "phone": link.guardian.phone or "---",
                "relation": link.relation or "---",
                "start_date": _format_date(link.start_date),
                "end_date": _format_date(link.end_date) if link.end_date else "مستمر",
                "is_primary": link.is_primary
            }
            guardian_history.append(g_data)
            if link.is_primary:
                primary_guardian = g_data

        return {
            **common,
            "orphan": {"name": orphan.name, "national_id": orphan.national_id or "---", "birth_date": _format_date(orphan.date_birth), "gender": gender_display, "phone": orphan.phone or '---'},
            "deceased": deceased_info,
            "primary_guardian": primary_guardian,
            "guardian_history": guardian_history,
            "balances": balances
        }

    return common


def fetch_deceased_report_data(deceased_id: int, db_service: DBService, user):
    deceased, orphans, _ = db_service.get_deceased_details(deceased_id)
    if not deceased: return None

    all_currencies = db_service.get_currencies()
    current_balances = {b.currency_id: float(b.balance) for b in deceased.balances}
    full_balances_list = [{"currency_name": curr.name, "currency_code": curr.code, "balance": current_balances.get(curr.id, 0.0)} for curr in all_currencies]

    orphans_list = []
    for o in orphans:
        primary_link = next((link for link in o.guardian_links if link.is_primary), None)
        guardian = primary_link.guardian if primary_link else None
        orphans_list.append({
            "name": o.name, "national_id": o.national_id or "---",
            "birth_date": o.date_birth.strftime("%Y/%m/%d") if o.date_birth else "---",
            "guardian_name": guardian.name if guardian else "غير محدد", 
            "guardian_nid": guardian.national_id or "---" if guardian else "---",
            "guardian_start_date": primary_link.start_date.strftime("%Y/%m/%d") if primary_link and primary_link.start_date else "---",
            "balances": _build_full_balances(o.balances, all_currencies)
        })

    return {
        "generated_at": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "organization": "مؤسسة إدارة وتنمية أموال اليتامى",
        "exported_by": user.name,
        "logo_path": logo_path,
        "ps_logo_path": ps_logo_path,
        "deceased": {"name": deceased.name, "national_id": deceased.national_id or "---", "death_date": deceased.date_death.strftime("%Y/%m/%d") if deceased.date_death else "---", "account_number": deceased.account_number or "---", "archives_number": deceased.archives_number or "---", "balances": full_balances_list},
        "orphans": orphans_list,
    }


def fetch_guardian_report_data(guardian_id: int, db_service: DBService, user):
    guardian, orphans_data = db_service.get_guardian_details(guardian_id)
    if not guardian: return None

    orphans_under_care = []
    for orphan in orphans_data:
        link = next((l for l in orphan.guardian_links if l.guardian_id == guardian_id), None)
        if link:
            all_currencies = db_service.get_currencies()
            orphans_under_care.append({
                "name": orphan.name, "national_id": orphan.national_id or "---",
                "relation": link.relation if link else "غير محدد",
                "start_date": link.start_date.strftime("%Y/%m/%d") if link.start_date else "---",
                "end_date": link.end_date.strftime("%Y/%m/%d") if link.end_date else "مستمر",
                "balances": _build_full_balances(orphan.balances, all_currencies),
                "is_primary": 'نعم' if link.is_primary else 'لا',
            })

    return {
        "generated_at": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "organization": "مؤسسة إدارة وتنمية أموال اليتامى",
        "exported_by": user.name,
        "logo_path": logo_path,
        "ps_logo_path": ps_logo_path,
        "guardian": {"name": guardian.name, "national_id": guardian.national_id or "---", "phone": guardian.phone or "---"},
        "orphans": orphans_under_care
    }


def fetch_monthly_orphans_report(from_date_str, to_date_str, db_service, current_user):
    from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
    to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
    orphans = db_service.get_orphans_by_date_range(from_date, to_date)
    all_currencies = db_service.get_currencies()
    return {
        "from_date": from_date_str, "to_date": to_date_str,
        "generated_at": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "orphans": [{"name": o.name, "national_id": o.national_id or "---", "balances": _build_full_balances(o.balances, all_currencies)} for o in orphans],
        "total_count": len(orphans),
        "exported_by": current_user
    }


def generate_monthly_report(from_date_str, to_date_str, output_path, db_service, current_user_name, file_format="pdf"):
    try:
        start_dt = datetime.strptime(from_date_str, "%Y-%m-%d")
        end_dt = datetime.strptime(to_date_str, "%Y-%m-%d")
        orphans = db_service.get_orphans_by_date_range(start_dt, end_dt)
        if not orphans:
            raise ValueError("لا توجد سجلات أيتام في هذه الفترة.")

        data_orphans = []
        all_currencies = db_service.get_currencies()
        for o in orphans:
            primary_link = next((link for link in o.guardian_links if link.is_primary), None)
            guardian_name = primary_link.guardian.name if primary_link else "غير محدد"
            data_orphans.append({
                "name": o.name, "national_id": o.national_id or "---",
                "birth_date": o.date_birth.strftime("%Y/%m/%d") if o.date_birth else "---",
                "gender": "ذكر" if o.gender.name == "male" else "أنثى",
                "age": calculate_age(o.date_birth) if o.date_birth else "---",
                "guardian": guardian_name,
                "balances": _build_full_balances(o.balances, all_currencies)
            })

        data = {
            "from_date": from_date_str, "to_date": to_date_str,
            "generated_at": datetime.now().strftime("%Y/%m/%d %H:%M"),
            "orphans": data_orphans, "total_count": len(orphans),
            "organization": "مؤسسة إدارة وتنمية أموال اليتامى",
            "exported_by": current_user_name,
            "logo_path": logo_path,
            "ps_logo_path": ps_logo_path,
        }

        if file_format == "pdf":
            return _export_as_pdf(data, output_path)
        else:
            return _export_as_excel(data, output_path)
    except Exception as e:
        print(f"Error in generation: {e}")
        raise e


def _export_as_pdf(data, output_path):
    template = env.get_template("monthly_report.html")
    html_content = template.render(**data)
    HTML(string=html_content).write_pdf(output_path)
    return True


def _export_as_excel(data, output_path):
    excel_data = []
    for i, o in enumerate(data['orphans'], 1):
        bal_text = " / ".join([f"{b['amount']} {b['currency']}" for b in o['balances']])
        excel_data.append({"م": i, "اسم اليتيم": o['name'], "الجنس": o['gender'], "تاريخ الميلاد": o['birth_date'], "العمر": o['age'], "الوصي الحالي": o['guardian'], "الأرصدة": bal_text})
    df = pd.DataFrame(excel_data)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='تقرير دوري')
        worksheet = writer.sheets['تقرير دوري']
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column].width = max_length + 5
    return True


def _format_money(value):
    try:
        return f"{Decimal(str(value or 0)):,.2f}"
    except Exception:
        return "0.00"


def _txn_datetime(txn):
    return getattr(txn, "created_at", None) or getattr(txn, "created_date", None)


def _build_financial_table_report_data(deceased_id: int, currency_id: int, db_service: DBService, exported_by: str):
    db = db_service.session
    deceased = db.query(Deceased).filter_by(id=deceased_id).first()
    if not deceased:
        raise ValueError("لم يتم العثور على المتوفى المطلوب.")

    if not currency_id:
        raise ValueError("يرجى اختيار العملة قبل تصدير التقرير.")

    currency = db.query(Currency).filter_by(id=currency_id).first()
    currency_label = f"{currency.name} ({currency.code})" if currency else str(currency_id)

    orphans = list(deceased.orphans or [])
    orphan_ids = [o.id for o in orphans]

    primary_guardian = None
    guardian_relation = ""
    if orphan_ids:
        primary_link = db.query(OrphanGuardian).filter(
            OrphanGuardian.orphan_id.in_(orphan_ids),
            OrphanGuardian.is_primary == True,
        ).first()
        if not primary_link:
            primary_link = db.query(OrphanGuardian).filter(
                OrphanGuardian.orphan_id.in_(orphan_ids)
            ).first()
        if primary_link:
            primary_guardian = primary_link.guardian
            guardian_relation = (primary_link.relation or "").strip()

    entities = []
    for orphan in orphans:
        display_name = (orphan.name or "").strip().split()[0] if orphan.name else ""
        entities.append({
            "kind": "orphan",
            "id": orphan.id,
            "name": orphan.name,
            "header": display_name or orphan.name or f"يتيم {orphan.id}",
        })

    if primary_guardian:
        g_name = (primary_guardian.name or "").strip().split()[0] if primary_guardian.name else ""
        g_header = g_name or primary_guardian.name or "الوصي الأساسي"
        if guardian_relation:
            g_header = f"{g_header} ({guardian_relation})"
        entities.append({
            "kind": "guardian",
            "id": primary_guardian.id,
            "name": primary_guardian.name,
            "header": g_header,
        })

    history_rows = []
    if orphan_ids:
        orphan_txns = db.query(Transaction).filter(
            Transaction.orphan_id.in_(orphan_ids),
            Transaction.currency_id == currency_id,
        ).order_by(Transaction.created_at).all()
        for txn in orphan_txns:
            history_rows.append({"kind": "orphan", "person_id": txn.orphan_id, "txn": txn})

    if primary_guardian:
        guardian_txns = db.query(GuardianTransaction).filter(
            GuardianTransaction.guardian_id == primary_guardian.id,
            GuardianTransaction.deceased_id == deceased.id,
            GuardianTransaction.currency_id == currency_id,
        ).order_by(GuardianTransaction.created_date).all()
        for txn in guardian_txns:
            history_rows.append({"kind": "guardian", "person_id": txn.guardian_id, "txn": txn})

    history_rows.sort(key=lambda rec: _txn_datetime(rec["txn"]) or datetime.min)

    grouped = {}
    for rec in history_rows:
        txn = rec["txn"]
        group_key = (getattr(txn, "row_group_key", None) or "").strip()
        dt = _txn_datetime(txn)
        if not group_key:
            if not dt:
                continue
            group_key = dt.strftime("%Y-%m-%d %H:%M:%S.%f")
        grouped.setdefault(group_key, []).append(rec)

    group_items = list(grouped.items())
    group_items.sort(key=lambda pair: _txn_datetime(pair[1][0]["txn"]) or datetime.min)

    rows = []
    for _, tx_list in group_items:
        first_txn = tx_list[0]["txn"]
        first_dt = _txn_datetime(first_txn)
        date_text = first_dt.strftime("%d/%m/%Y") if first_dt else "---"

        sums = {}
        for rec in tx_list:
            txn = rec["txn"]
            key = (rec["kind"], rec["person_id"])
            sums.setdefault(key, {"deposit": Decimal("0"), "withdraw": Decimal("0")})
            amount = Decimal(str(txn.amount or 0))
            if txn.type == TransactionTypeEnum.deposit:
                sums[key]["deposit"] += amount
            else:
                sums[key]["withdraw"] += amount

        row_cells = []
        total_balance = Decimal("0")
        for entity in entities:
            key = (entity["kind"], entity["id"])
            values = sums.get(key, {"deposit": Decimal("0"), "withdraw": Decimal("0")})
            dep = values["deposit"]
            wd = values["withdraw"]
            total_balance += dep - wd

            if dep > 0 and wd > 0:
                cell_text = f"إيداع {_format_money(dep)} | سحب {_format_money(wd)}"
            elif dep > 0:
                cell_text = f"{_format_money(dep)}"
            elif wd > 0:
                cell_text = f"-{_format_money(wd)}"
            else:
                cell_text = ""
            row_cells.append(cell_text)

        rows.append({
            "date": date_text,
            "cells": row_cells,
            "total_balance": _format_money(total_balance),
            "note": str(getattr(first_txn, "note", "") or "").strip(),
        })

    deceased_deposit = db.query(DeceasedTransaction).filter(
        DeceasedTransaction.deceased_id == deceased.id,
        DeceasedTransaction.currency_id == currency_id,
        DeceasedTransaction.type == TransactionTypeEnum.deposit,
    ).all()
    deceased_withdraw = db.query(DeceasedTransaction).filter(
        DeceasedTransaction.deceased_id == deceased.id,
        DeceasedTransaction.currency_id == currency_id,
        DeceasedTransaction.type == TransactionTypeEnum.withdraw,
    ).all()

    total_deposited = sum(Decimal(str(tx.amount or 0)) for tx in deceased_deposit)
    total_withdrawn = sum(Decimal(str(tx.amount or 0)) for tx in deceased_withdraw)

    bal = db.query(DeceasedBalance).filter_by(
        deceased_id=deceased.id,
        currency_id=currency_id,
    ).first()
    available_balance = Decimal(str(bal.balance or 0)) if bal else Decimal("0")

    return {
        "generated_at": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "organization": "مؤسسة إدارة وتنمية أموال اليتامى",
        "exported_by": exported_by,
        "logo_path": logo_path,
        "ps_logo_path": ps_logo_path,
        "deceased": {
            "id": deceased.id,
            "name": deceased.name,
            "national_id": deceased.national_id or "---",
            "archives_number": deceased.archives_number or "---",
            "currency_label": currency_label,
            "total_deposited": _format_money(total_deposited),
            "total_withdrawn": _format_money(total_withdrawn),
            "available_balance": _format_money(available_balance),
        },
        "entity_headers": [entity["header"] for entity in entities],
        "rows": rows,
    }


def _export_financial_table_excel(data, output_path):
    headers = ["تاريخ الحركة", *data.get("entity_headers", []), "الرصيد الكلي", "ملاحظة"]

    records = []
    for row in data.get("rows", []):
        records.append([row.get("date", ""), *row.get("cells", []), row.get("total_balance", "0.00"), row.get("note", "")])

    if not records:
        records = [["---", *("" for _ in data.get("entity_headers", [])), "0.00", ""]]

    df = pd.DataFrame(records, columns=headers)
    summary_df = pd.DataFrame([
        ["اسم المتوفى", data["deceased"]["name"]],
        ["رقم الهوية", data["deceased"]["national_id"]],
        ["رقم الأرشيف", data["deceased"]["archives_number"]],
        ["العملة", data["deceased"]["currency_label"]],
        ["إجمالي المودع", data["deceased"]["total_deposited"]],
        ["إجمالي المسحوب", data["deceased"]["total_withdrawn"]],
        ["إجمالي المتاح", data["deceased"]["available_balance"]],
    ], columns=["البيان", "القيمة"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="الملخص")
        df.to_excel(writer, index=False, sheet_name="حركات المتوفى")

        summary_ws = writer.sheets["الملخص"]
        table_ws = writer.sheets["حركات المتوفى"]

        header_fill = PatternFill(start_color="1A2A6C", end_color="1A2A6C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for ws in (summary_ws, table_ws):
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[letter].width = min(max(max_len + 4, 14), 42)

    return True


def generate_financial_table_report(
    deceased_id: int,
    currency_id: int,
    output_path: str,
    db_service: DBService,
    exported_by: str,
    file_format: str = "pdf",
):
    data = _build_financial_table_report_data(
        deceased_id=deceased_id,
        currency_id=currency_id,
        db_service=db_service,
        exported_by=exported_by,
    )

    if str(file_format).lower() == "excel":
        return _export_financial_table_excel(data, output_path)

    template = env.get_template("financial_table_report.html")
    html_content = template.render(**data)
    if RENDERER == "weasy":
        HTML(string=html_content).write_pdf(output_path)
    else:
        pdfkit.from_string(html_content, output_path)
    return True


def generate_report(entity_type: str, entity_id: int, output_path: str, user):
    db = DBService()
    try:
        if entity_type == "deceased":
            data = fetch_deceased_report_data(entity_id, db, user)
            template_name = "deceased_report.html"
        elif entity_type == "orphan":
            data = fetch_entity_data("orphan", entity_id, db, user)
            template_name = "orphan_report.html"
        elif entity_type == "guardian":
            data = fetch_guardian_report_data(entity_id, db, user)
            template_name = "guardian_report.html"
        else:
            raise ValueError("نوع التقرير غير مدعوم حالياً")

        if not data:
            raise Exception("لم يتم العثور على بيانات للسجل المطلوب")

        template = env.get_template(template_name)
        html_content = template.render(**data)
        if RENDERER == "weasy":
            HTML(string=html_content).write_pdf(output_path)
        else:
            pdfkit.from_string(html_content, output_path)
        return output_path
    finally:
        db.close()
